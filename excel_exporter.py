import sqlite3
from datetime import datetime, timedelta, time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

REPORT_FOLDER = "/mnt/rdrive/Share/ProductionData"
EXCEL_FILE = f"{REPORT_FOLDER}/production_report.xlsx"
DB_FILE = "production.db"
EXPORT_STATUS_FILE = f"{REPORT_FOLDER}/last_export.txt"

conn = sqlite3.connect(DB_FILE)
cursor = conn.cursor()

# Lista masinilor
cursor.execute("""
SELECT DISTINCT machine
FROM production
ORDER BY machine
""")

machines = [row[0] for row in cursor.fetchall()]

# ==================================================
# SHIFT DEFINITIONS
# ==================================================

SHIFT_1_START = time(6, 0)
SHIFT_1_END = time(14, 0)

SHIFT_2_START = time(14, 0)
SHIFT_2_END = time(22, 0)


def get_shift_from_datetime(dt):

    t = dt.time()

    if SHIFT_1_START <= t < SHIFT_1_END:
        return 1

    if SHIFT_2_START <= t < SHIFT_2_END:
        return 2

    return None


# ==================================================
# SPLIT EVENT INTO SHIFTS
# ==================================================

def split_event_into_shifts(start, end):

    result = []

    current_day = start.date()

    while current_day <= end.date():

        shift1_start = datetime.combine(
            current_day,
            SHIFT_1_START
        )

        shift1_end = datetime.combine(
            current_day,
            SHIFT_1_END
        )

        shift2_start = datetime.combine(
            current_day,
            SHIFT_2_START
        )

        shift2_end = datetime.combine(
            current_day,
            SHIFT_2_END
        )

        # ----------------------------
        # SHIFT 1
        # ----------------------------

        s = max(start, shift1_start)
        e = min(end, shift1_end)

        if s < e:

            result.append(
                (
                    current_day.isoformat(),
                    1,
                    int(
                        (e - s).total_seconds()
                    )
                )
            )

        # ----------------------------
        # SHIFT 2
        # ----------------------------

        s = max(start, shift2_start)
        e = min(end, shift2_end)

        if s < e:

            result.append(
                (
                    current_day.isoformat(),
                    2,
                    int(
                        (e - s).total_seconds()
                    )
                )
            )

        current_day += timedelta(days=1)

    return result

wb = Workbook()

# Primul sheet va fi Machine Status
ws_status = wb.active
ws_status.title = "Machine Status"

# =====================================
# MACHINE STATUS
# =====================================

ws_status.append([
    "Date",
    "Shift",
    "Machine",
    "Pieces",
    "Running (min)",
    "Fault (min)",
    "Stopped (min)",
    "Utilization (%)",
    "Pieces/Hour"
])

for cell in ws_status[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

ws_status.freeze_panes = "A2"
ws_status.auto_filter.ref = "A1:I1"

for col in "ABCDEFGHI":
    ws_status.column_dimensions[col].width = 18


# ==================================================
# PIECES
# ==================================================

cursor.execute("""
SELECT
    DATE(timestamp),
    shift,
    machine,
    COUNT(*)
FROM production
GROUP BY
    DATE(timestamp),
    shift,
    machine
""")

piece_counts = {}

for date, shift, machine, pieces in cursor.fetchall():

    piece_counts[
        (date, shift, machine)
    ] = pieces


# ==================================================
# STATES
# ==================================================

cursor.execute("""
SELECT
    machine,
    state,
    start_time,
    end_time
FROM machine_state_events
ORDER BY start_time
""")

status_rows = {}

for machine, state, start_str, end_str in cursor.fetchall():

    start = datetime.strptime(
        start_str,
        "%Y-%m-%d %H:%M:%S"
    )

    end = datetime.strptime(
        end_str,
        "%Y-%m-%d %H:%M:%S"
    )

    parts = split_event_into_shifts(
        start,
        end
    )

    for date, shift, duration in parts:

        key = (
            date,
            shift,
            machine
        )

        if key not in status_rows:

            status_rows[key] = {
                "RUNNING": 0,
                "FAULT": 0,
                "STOPPED": 0
            }

        status_rows[key][state] += duration


# ==================================================
# WRITE REPORT
# ==================================================

for key in sorted(status_rows):

    date, shift, machine = key

    values = status_rows[key]

    running = values["RUNNING"] / 60
    fault = values["FAULT"] / 60
    stopped = values["STOPPED"] / 60

    pieces = piece_counts.get(
        key,
        0
    )

    utilization = (
        running / 480
    ) * 100

    if running > 0:

        pieces_per_hour = (
            pieces /
            (running / 60)
        )

    else:

        pieces_per_hour = 0

    ws_status.append([
        date,
        shift,
        machine,
        pieces,
        round(running, 1),
        round(fault, 1),
        round(stopped, 1),
        round(utilization, 1),
        round(pieces_per_hour, 1)
    ])

# =====================================
# SHEET PER MASINA
# =====================================

for machine in machines:

    ws = wb.create_sheet(machine)

    ws.append([
        "Date",
        "Interval",
        "Shift",
        "Pieces"
    ])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12

    cursor.execute("""
    SELECT
        DATE(timestamp),
        STRFTIME('%H', timestamp),
        shift,
        COUNT(*)
    FROM production
    WHERE machine = ?
    GROUP BY
        DATE(timestamp),
        STRFTIME('%H', timestamp),
        shift
    ORDER BY
        DATE(timestamp),
        STRFTIME('%H', timestamp)
    """, (machine,))

    for row in cursor.fetchall():
        row = list(row)
        hour = int(row[1])
        row[1] = f"{hour:02d} - {hour+1:02d}"
        ws.append(row)

    for cell in ws["B"][1:]:
        cell.number_format = "00"

wb.save(EXCEL_FILE)

from datetime import datetime

with open(EXPORT_STATUS_FILE, "w") as f:

    f.write(
        datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )
    )

conn.close()

print(f"Created {EXCEL_FILE}")

